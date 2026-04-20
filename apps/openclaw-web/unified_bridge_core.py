"""
Unified Bridge Core — MiniMax API 路由层
route_text / route_image / route_generic_file 统一接入 MiniMax M2.7
"""
import os, requests, json, base64, tempfile
from pathlib import Path

MINIMAX_API_KEY = os.getenv("MINIMAX_API_KEY", "")
MINIMAX_URL     = os.getenv("MINIMAX_URL", "https://api.minimaxi.com/anthropic/v1/messages")
MODEL           = os.getenv("MINIMAX_MODEL", "MiniMax-M2.7")

_sessions = {}

def _get_history(session_id: str) -> list[dict]:
    if session_id not in _sessions:
        _sessions[session_id] = []
    return _sessions[session_id]

def _call_minimax(messages: list[dict], session_id: str = "default") -> str:
    if not MINIMAX_API_KEY:
        return "[配置错误] MINIMAX_API_KEY 未配置"
    history = _get_history(session_id)
    context = history[-10:] if history else []
    full_messages = context + messages

    headers = {
        "Authorization": f"Bearer {MINIMAX_API_KEY}",
        "Content-Type": "application/json",
        "anthropic-version": "2023-06-01"
    }
    payload = {
        "model": MODEL,
        "messages": full_messages,
        "max_tokens": 4096,
        "temperature": 0.7,
    }

    try:
        resp = requests.post(MINIMAX_URL, headers=headers, json=payload, timeout=90)
        if resp.status_code != 200:
            return f"[API错误 {resp.status_code}] {resp.text[:200]}"
        data = resp.json()
        history.append(messages[-1])
        answer = ""
        for block in data.get("content", []):
            if block.get("type") == "text" and block.get("text"):
                answer += block["text"]
        history.append({"role": "assistant", "content": answer})
        return answer
    except Exception as e:
        return f"[网络错误] {e}"

def _upload_to_data_url(file_path: str, max_size_mb: int = 3) -> str:
    path = Path(file_path)
    size_mb = path.stat().st_size / 1024 / 1024
    if size_mb > max_size_mb:
        with open(file_path, "rb") as f:
            data = f.read(int(max_size_mb * 1024 * 1024))
    else:
        with open(file_path, "rb") as f:
            data = f.read()
    ext = path.suffix.lower()
    mime_map = {".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".png": "image/png",
                ".gif": "image/gif", ".webp": "image/webp", ".bmp": "image/bmp"}
    mime = mime_map.get(ext, "application/octet-stream")
    b64 = base64.b64encode(data).decode()
    return f"data:{mime};base64,{b64}"

def _extract_file_text(file_path: str, original_name: str = "") -> str:
    """从各类文件提取文本内容"""
    path = Path(file_path)
    ext = (original_name or path.name).lower()

    try:
        # TXT
        if ext.endswith((".txt", ".md", ".json", ".xml", ".html", ".log", ".yaml", ".yml", ".py", ".js")):
            for enc in ("utf-8", "gbk", "gb2312", "latin1"):
                try:
                    return Path(file_path).read_text(encoding=enc)
                except:
                    continue

        # PDF
        if ext.endswith(".pdf") or not ext:
            try:
                import PyPDF2
                reader = PyPDF2.PdfReader(file_path)
                texts = []
                for page in reader.pages[:15]:
                    t = page.extract_text()
                    if t:
                        texts.append(t[:2000])
                if texts:
                    return f"[PDF内容]\n" + "\n\n".join(texts)
            except:
                pass

        # DOCX
        if ext.endswith(".docx"):
            try:
                from docx import Document
                doc = Document(file_path)
                return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
            except:
                pass

        # XLSX
        if ext.endswith(".xlsx"):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(file_path, data_only=True)
                parts = []
                for sheet in wb.sheetnames[:3]:
                    ws = wb[sheet]
                    rows = []
                    for row in ws.iter_rows(max_row=100, values_only=True):
                        if any(c is not None for c in row):
                            rows.append(" | ".join(str(c) if c is not None else "" for c in row))
                    if rows:
                        parts.append(f"[Sheet: {sheet}]\n" + "\n".join(rows[:50]))
                return "\n\n".join(parts)
            except:
                pass

    except Exception as e:
        return f"[文件解析错误: {e}]"

    return ""

def route_text(text: str, session_id: str = "default") -> str:
    """纯文本消息路由"""
    return _call_minimax([{"role": "user", "content": text}], session_id=session_id)

def route_image(file_path: str, caption: str = "", session_id: str = "default") -> str:
    """图片消息路由"""
    paths = [p.strip() for p in file_path.split("|") if p.strip()]
    content = []
    for p in paths:
        data_url = _upload_to_data_url(p)
        mime = data_url.split(";")[0].split(":")[1]
        b64_data = data_url.split(",", 1)[1]
        content.append({
            "type": "image",
            "source": {"type": "base64", "media_type": mime, "data": b64_data}
        })
    if caption:
        content.append({"type": "text", "text": caption})
    return _call_minimax([{"role": "user", "content": content}], session_id=session_id)

def route_generic_file(file_path: str, original_name: str = "", caption: str = "", session_id: str = "default") -> str:
    """通用文件路由"""
    text_content = _extract_file_text(file_path, original_name)
    if text_content:
        prompt = f"【附件内容】\n{text_content[:8000]}"
        if caption:
            prompt += f"\n\n【用户说明】{caption}"
    elif caption:
        prompt = caption
    else:
        return "[文件处理失败] 无法提取附件内容，请重试或换一种格式。"
    return _call_minimax([{"role": "user", "content": prompt}], session_id=session_id)
